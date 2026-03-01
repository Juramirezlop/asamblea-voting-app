import logging
import re
import pandas as pd
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from ..database import get_db, execute_query, close_db
from ..auth.auth import admin_required
from typing import Dict
from fastapi.responses import StreamingResponse
from fpdf import FPDF
from io import BytesIO
from pydantic import BaseModel
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

router = APIRouter(prefix="/participants", tags=["Participants"])
logger = logging.getLogger(__name__)

class ConjuntoRequest(BaseModel):
    nombre: str

# Listar participantes (solo admin)
@router.get("/", dependencies=[Depends(admin_required)])
def listar_participantes():
    conn = get_db()
    try:
        participants = execute_query(conn, "SELECT * FROM participants", fetchall=True)
        return [dict(p) for p in participants]
    finally:
        close_db(conn)

# Guardar nombre del conjunto
@router.post("/conjunto/nombre", dependencies=[Depends(admin_required)])
async def guardar_nombre_conjunto(request: ConjuntoRequest):
    conn = get_db()
    try:
        execute_query(
            conn,
            "INSERT INTO config (key, value) VALUES (?, ?) ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value",
            ("conjunto_nombre", request.nombre),
            commit=True
        )
        from ..main import manager
        await manager.broadcast_to_admins({
            "type": "conjunto_name_updated",
            "data": {"nombre": request.nombre}
        })
        return {"status": "ok"}
    finally:
        close_db(conn)

# Obtener nombre conjunto
@router.get("/conjunto/nombre", dependencies=[Depends(admin_required)])
def obtener_nombre_conjunto():
    conn = get_db()
    try:
        result = execute_query(conn, "SELECT value FROM config WHERE key = ?", ("conjunto_nombre",), fetchone=True)
        return {"nombre": result["value"] if result and result["value"] else None}
    finally:
        close_db(conn)

# Obtener nombre conjunto (público)
@router.get("/conjunto/nombre/public")
def obtener_nombre_conjunto_publico():
    conn = get_db()
    try:
        result = execute_query(conn, "SELECT value FROM config WHERE key = ?", ("conjunto_nombre",), fetchone=True)
        return {"nombre": result["value"] if result and result["value"] else "Conjunto Residencial"}
    finally:
        close_db(conn)

# Carga masiva desde JSON
@router.post("/bulk", dependencies=[Depends(admin_required)])
async def agregar_participantes(data: Dict[str, dict]):
    conn = get_db()
    count = 0
    try:
        for code, info in data.items():
            name = info.get("nombre") or info.get("name")
            coef = info.get("coeficiente") or info.get("coefficient") or 1.0
            ha_votado = int(bool(info.get("ha_votado", False)))
            if not code or not name:
                continue
            execute_query(
                conn,
                """
                INSERT INTO participants (code, name, coefficient, has_voted, present)
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT (code) DO UPDATE SET 
                    name = EXCLUDED.name,
                    coefficient = EXCLUDED.coefficient,
                    has_voted = EXCLUDED.has_voted,
                    present = EXCLUDED.present
                """,
                (code.upper(), name, float(coef), ha_votado, 0),
                commit=True
            )
            count += 1
        from ..main import manager
        await manager.broadcast_to_admins({"type": "participants_bulk_loaded", "data": {"count": count}})
        return {"status": "ok", "inserted": count, "message": f"✅ {count} participantes cargados exitosamente", "total_participants": count}
    finally:
        close_db(conn)


def extraer_torre_id(df, sheet_name):
    """
    Extrae el identificador de torre desde la fila 2 del Excel.
    Soporta letras (A, B, C, L...) y números (1, 2, 3...).
    Es case-insensitive y robusto ante distintos formatos.
    """
    torre_id = None

    if len(df) > 1:
        for col in df.columns:
            try:
                cell_value = str(df.iloc[1, col]).strip().upper()
            except Exception:
                continue

            # "TORRE A", "TORRE B", "TORRE L", etc.
            m = re.search(r'TORRE\s+([A-Z])(?:\s|$)', cell_value)
            if m:
                return m.group(1)

            # "TORRE 1", "TORRE NÚMERO 2", etc.
            m = re.search(r'TORRE\s+(?:N[ÚU]MERO\s+)?(\d+)', cell_value)
            if m:
                return m.group(1)

            # Valor corto alfanumérico suelto (ej. "A", "B", "1", "L")
            m = re.match(r'^([A-Z0-9]{1,3})$', cell_value)
            if m:
                torre_id = m.group(1)

    if torre_id:
        return torre_id

    # Fallback: extraer del nombre de la hoja
    m = re.search(r'TORRE\s+([A-Z])(?:\s|$)', sheet_name.upper())
    if m:
        return m.group(1)

    m = re.search(r'(?:^|\s)([A-Z])(?:\s|$)', sheet_name.upper())
    if m:
        return m.group(1)

    m = re.search(r'(\d+)', sheet_name)
    if m:
        return m.group(1)

    return "1"  # Fallback final


# Endpoint para subir un XLSX desde admin
@router.post("/upload-xlsx", dependencies=[Depends(admin_required)])
async def upload_xlsx(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        xls = pd.read_excel(BytesIO(contents), sheet_name=None, header=None)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error leyendo Excel: {e}")

    participantes = {}

    for sheet_name, df in xls.items():
        # Detectar torre_id (soporta letras y números, case-insensitive)
        torre_id = extraer_torre_id(df, sheet_name)
        logger.info(f"Hoja '{sheet_name}' -> torre_id='{torre_id}'")

        try:
            df_data = pd.read_excel(BytesIO(contents), sheet_name=sheet_name, header=2)
        except Exception as e:
            logger.warning(f"No se pudo leer hoja '{sheet_name}': {e}")
            continue

        df_data.columns = df_data.columns.str.strip().str.upper()
        required_cols = {"NO APTO", "PROPIETARIO", "COEFICIENTE"}
        if not required_cols.issubset(set(df_data.columns)):
            logger.warning(f"Hoja '{sheet_name}' no tiene columnas requeridas. Columnas encontradas: {list(df_data.columns)}")
            continue

        for _, row in df_data.iterrows():
            try:
                apto = str(row["NO APTO"]).strip()
                nombre = str(row["PROPIETARIO"]).strip()
                coef_str = str(row["COEFICIENTE"]).replace(",", ".").strip()
                coef = float(coef_str) if coef_str not in ("", "nan", "None") else 1.0

                if not apto or not nombre or apto == "nan" or nombre == "nan":
                    continue

                codigo = f"{torre_id}-{apto}"
                participantes[codigo] = {
                    "nombre": nombre,
                    "coeficiente": coef,
                    "ha_votado": False
                }
            except Exception:
                continue

    # Insertar en DB
    conn = get_db()
    inserted = 0
    try:
        for code, info in participantes.items():
            execute_query(
                conn,
                """
                INSERT INTO participants (code, name, coefficient, has_voted, present)
                VALUES (?, ?, ?, ?, 0)
                ON CONFLICT (code) DO UPDATE SET
                    name = EXCLUDED.name,
                    coefficient = EXCLUDED.coefficient,
                    has_voted = EXCLUDED.has_voted
                """,
                (code.upper(), info["nombre"], float(info["coeficiente"]), int(info.get("ha_votado", False))),
                commit=True
            )
            inserted += 1

        from ..main import manager
        await manager.broadcast_to_admins({
            "type": "excel_uploaded",
            "data": {"inserted": inserted, "sheets_processed": len(xls.keys())}
        })

        return {"status": "ok", "inserted": inserted, "sheets_processed": len(xls.keys())}
    finally:
        close_db(conn)


# Genera PDF de asistencia
@router.post("/asistencia/pdf")
async def generar_pdf_asistencia(user=Depends(admin_required)):
    colombia_tz = timezone(timedelta(hours=-5))
    fecha_actual = datetime.now(colombia_tz)
    conn = get_db()

    try:
        conjunto_result = execute_query(conn, "SELECT value FROM config WHERE key = ?", ("conjunto_nombre",), fetchone=True)
        conjunto_name = conjunto_result["value"] if conjunto_result and conjunto_result.get("value") else "Conjunto Residencial"

        participantes = execute_query(
            conn,
            "SELECT code, name, coefficient, present, is_power, login_time FROM participants ORDER BY code",
            fetchall=True
        )

        stats = execute_query(
            conn,
            """
            SELECT 
                COUNT(*) as total_participants,
                COUNT(CASE WHEN present = 1 THEN 1 END) as present_count,
                COUNT(CASE WHEN present = 1 AND is_power = FALSE THEN 1 END) as own_votes,
                COUNT(CASE WHEN present = 1 AND is_power = TRUE THEN 1 END) as power_votes,
                COALESCE(SUM(CASE WHEN present = 1 THEN coefficient ELSE 0 END), 0) as present_coefficient,
                COALESCE(SUM(coefficient), 0) as total_coefficient
            FROM participants
            """,
            fetchone=True
        )

        if not stats:
            raise HTTPException(status_code=500, detail="Error obteniendo estadísticas")

        coefficient_percentage = float(stats['present_coefficient']) if stats['present_coefficient'] else 0.0
        quorum_met = coefficient_percentage >= 51

        preguntas = execute_query(
            conn,
            "SELECT DISTINCT q.id, q.text, q.type, q.allow_multiple, q.max_selections FROM questions q ORDER BY q.id",
            fetchall=True
        )

        resultados_preguntas = []
        if preguntas:
            for pregunta in preguntas:
                try:
                    opciones_result = execute_query(
                        conn,
                        "SELECT option_text FROM options WHERE question_id = ? ORDER BY option_text",
                        (pregunta['id'],),
                        fetchall=True
                    )
                    opciones = [r["option_text"] for r in opciones_result] if opciones_result else []

                    total_participants_result = execute_query(
                        conn,
                        "SELECT COUNT(DISTINCT participant_code) as total_participants FROM votes WHERE question_id = ?",
                        (pregunta['id'],),
                        fetchone=True
                    )
                    total_participants_pregunta = int(total_participants_result['total_participants']) if total_participants_result and total_participants_result['total_participants'] else 0

                    total_coef_result = execute_query(
                        conn,
                        """
                        SELECT COALESCE(SUM(p.coefficient), 0) as total_participant_coefficient 
                        FROM (SELECT DISTINCT participant_code FROM votes WHERE question_id = ?) v
                        JOIN participants p ON v.participant_code = p.code
                        """,
                        (pregunta['id'],),
                        fetchone=True
                    )
                    total_participant_coefficient = float(total_coef_result["total_participant_coefficient"]) if total_coef_result else 0.0

                    resultados = []
                    for opcion in opciones:
                        result = execute_query(
                            conn,
                            """
                            SELECT 
                                COUNT(DISTINCT v.participant_code) as unique_voters,
                                COALESCE(SUM(DISTINCT p.coefficient), 0) as coefficient_sum
                            FROM votes v
                            JOIN participants p ON v.participant_code = p.code
                            WHERE v.question_id = ? 
                            AND (v.answer = ? OR v.answer LIKE ? OR v.answer LIKE ? OR v.answer LIKE ?)
                            """,
                            (pregunta['id'], opcion, f"{opcion},%", f"%, {opcion},%", f"%, {opcion}"),
                            fetchone=True
                        )
                        votes = int(result['unique_voters']) if result and result['unique_voters'] else 0
                        coefficient_sum = float(result['coefficient_sum']) if result and result['coefficient_sum'] else 0.0
                        resultados.append({'answer': opcion, 'votes': votes, 'coefficient_sum': coefficient_sum})

                    resultados.sort(key=lambda x: x['coefficient_sum'], reverse=True)
                    resultados_preguntas.append({
                        'pregunta': pregunta,
                        'resultados': resultados,
                        'total_participants': total_participants_pregunta,
                        'total_participant_coefficient': total_participant_coefficient
                    })
                except Exception as e:
                    logger.error(f"Error procesando pregunta {pregunta['id']}: {e}")
                    continue

    except Exception as e:
        logger.error(f"Error en generar_pdf_asistencia: {e}")
        raise HTTPException(status_code=500, detail=f"Error generando PDF: {str(e)}")
    finally:
        close_db(conn)

    try:
        pdf = FPDF()
        pdf.add_page()

        pdf.set_font("Helvetica", 'B', 18)
        pdf.cell(0, 12, "REPORTE COMPLETO DE ASAMBLEA", ln=True, align="C")
        pdf.set_font("Helvetica", 'B', 14)
        pdf.cell(0, 8, conjunto_name, ln=True, align="C")
        pdf.set_font("Helvetica", size=10)
        pdf.cell(0, 6, f"Fecha: {fecha_actual.strftime('%d/%m/%Y %H:%M')}", ln=True, align="C")
        pdf.ln(8)

        pdf.set_font("Helvetica", 'B', 12)
        pdf.cell(0, 8, "1. LISTA DE ASISTENCIA DETALLADA", ln=True)
        pdf.ln(5)

        pdf.set_font("Helvetica", 'B', 8)
        pdf.cell(15, 8, "No.", border=1, align="C")
        pdf.cell(25, 8, "Apartamento", border=1, align="C")
        pdf.cell(50, 8, "Nombre", border=1, align="C")
        pdf.cell(20, 8, "Coeficiente", border=1, align="C")
        pdf.cell(30, 8, "Fecha Ingreso", border=1, align="C")
        pdf.cell(20, 8, "Asistencia", border=1, align="C")
        pdf.cell(20, 8, "Poder", border=1, align="C", ln=True)

        pdf.set_font("Helvetica", size=8)
        id_counter = 1
        for p in participantes:
            fecha_ingreso = "-"
            if p.get("login_time") and p.get("present"):
                try:
                    login_time_str = str(p["login_time"])
                    if 'T' in login_time_str:
                        dt_utc = datetime.fromisoformat(login_time_str.replace('Z', '+00:00'))
                    else:
                        dt_utc = datetime.strptime(login_time_str, '%Y-%m-%d %H:%M:%S').replace(tzinfo=timezone.utc)
                    fecha_ingreso = dt_utc.astimezone(colombia_tz).strftime('%d/%m/%Y %H:%M')
                except Exception:
                    fecha_ingreso = "Error"

            asistencia = "SÍ" if p.get("present") else "NO"
            poder = "SÍ" if p.get("is_power") else "NO"

            pdf.cell(15, 6, str(id_counter), border=1, align="C")
            pdf.cell(25, 6, str(p.get("code", "")), border=1, align="C")
            pdf.cell(50, 6, str(p.get("name", ""))[:25], border=1)
            pdf.cell(20, 6, f"{float(p.get('coefficient', 0)):.2f}", border=1, align="C")
            pdf.cell(30, 6, fecha_ingreso, border=1, align="C")
            pdf.cell(20, 6, asistencia, border=1, align="C")
            pdf.cell(20, 6, poder if p.get("present") else "-", border=1, align="C", ln=True)
            id_counter += 1

        pdf.add_page()
        pdf.set_font("Helvetica", 'B', 12)
        pdf.cell(0, 8, "2. ESTADÍSTICAS GENERALES", ln=True)
        pdf.ln(2)
        pdf.set_font("Helvetica", size=9)
        for stat in [
            f"Total participantes registrados: {stats['total_participants']}",
            f"Participantes presentes: {stats['present_count']}",
            f"Votos por apto propio: {stats['own_votes']}",
            f"Votos por poder: {stats['power_votes']}",
            f"Participación por coeficiente: {coefficient_percentage:.2f}%"
        ]:
            pdf.cell(0, 6, stat, ln=True)
        pdf.ln(3)

        pdf.set_font("Helvetica", 'B', 12)
        pdf.cell(0, 8, "3. ESTADO DEL QUÓRUM", ln=True)
        pdf.ln(2)
        if quorum_met:
            pdf.set_text_color(0, 128, 0)
            pdf.set_font("Helvetica", 'B', 11)
            pdf.cell(0, 8, f"QUÓRUM ALCANZADO ({coefficient_percentage:.2f}% >= 51%)", ln=True)
        else:
            pdf.set_text_color(255, 0, 0)
            pdf.set_font("Helvetica", 'B', 11)
            pdf.cell(0, 8, f"SIN QUÓRUM ({coefficient_percentage:.2f}% < 51%)", ln=True)
        pdf.set_text_color(0, 0, 0)
        pdf.ln(5)

        pdf.set_font("Helvetica", 'B', 12)
        pdf.cell(0, 8, "4. RESULTADOS DE VOTACIONES", ln=True)
        pdf.ln(3)

        if resultados_preguntas:
            for i, resultado in enumerate(resultados_preguntas, 1):
                pregunta = resultado['pregunta']
                resultados = resultado['resultados']

                pdf.set_font("Helvetica", 'B', 10)
                pdf.cell(0, 7, f"Pregunta {i}: {str(pregunta['text'])}", ln=True)
                pdf.ln(2)
                pdf.set_font("Helvetica", size=8)
                pdf.cell(0, 4, f"Total presentes en asamblea: {stats['present_count']}", ln=True)
                pdf.cell(0, 4, f"Participaron en esta votación: {resultado['total_participants']}", ln=True)
                pdf.cell(0, 4, f"Coeficiente total: {resultado['total_participant_coefficient']:.2f}", ln=True)
                pdf.ln(3)

                if resultados:
                    resultados_ordenados = sorted(resultados, key=lambda x: x['coefficient_sum'], reverse=True)
                    if pregunta['allow_multiple']:
                        ganadoras = {r['answer'] for r in resultados_ordenados[:pregunta['max_selections']]}
                    else:
                        max_coef = resultados_ordenados[0]['coefficient_sum']
                        ganadoras = {r['answer'] for r in resultados_ordenados if abs(r['coefficient_sum'] - max_coef) < 0.01}

                    pdf.set_font("Helvetica", size=8)
                    pdf.cell(0, 5, "Opciones:", ln=True)
                    for res in resultados_ordenados:
                        if res['answer'] in ganadoras:
                            pdf.set_text_color(0, 128, 0)
                            pdf.set_font("Helvetica", 'B', 8)
                        else:
                            pdf.set_text_color(0, 0, 0)
                            pdf.set_font("Helvetica", size=8)
                        pdf.cell(0, 5, f"- {res['answer']}: {res['coefficient_sum']:.2f} % ({res['votes']} votos)", ln=True)
                    pdf.set_text_color(0, 0, 0)
                    pdf.set_font("Helvetica", size=8)
                else:
                    pdf.cell(0, 6, "Sin votos registrados", ln=True)
                pdf.ln(8)
        else:
            pdf.set_font("Helvetica", size=10)
            pdf.cell(0, 6, "No se han realizado votaciones en esta asamblea.", ln=True)

        pdf_bytes = pdf.output(dest="S")
        if isinstance(pdf_bytes, str):
            pdf_bytes = pdf_bytes.encode('latin-1')

        buffer = BytesIO(pdf_bytes)
        buffer.seek(0)
        return StreamingResponse(buffer, media_type="application/pdf", headers={
            "Content-Disposition": f"attachment; filename=reporte_completo_{conjunto_name.replace(' ', '_')}.pdf"
        })

    except Exception as e:
        logger.error(f"Error creando PDF: {e}")
        raise HTTPException(status_code=500, detail=f"Error creando PDF: {str(e)}")


@router.post("/asistencia/xlsx")
async def generar_xlsx_asistencia(user=Depends(admin_required)):
    colombia_tz = timezone(timedelta(hours=-5))
    fecha_actual = datetime.now(colombia_tz)
    conn = get_db()

    try:
        conjunto_result = execute_query(conn, "SELECT value FROM config WHERE key = ?", ("conjunto_nombre",), fetchone=True)
        conjunto_name = conjunto_result["value"] if conjunto_result and conjunto_result.get("value") else "Conjunto Residencial"

        participantes = execute_query(
            conn,
            "SELECT code, name, coefficient, present, is_power, login_time FROM participants ORDER BY code",
            fetchall=True
        )
    except Exception as e:
        logger.error(f"Error en generar_xlsx_asistencia: {e}")
        raise HTTPException(status_code=500, detail=f"Error obteniendo datos: {str(e)}")
    finally:
        close_db(conn)

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Asistencia"
        ws['A1'] = f"LISTA DE ASISTENCIA - {conjunto_name}"
        ws['A2'] = f"Fecha: {fecha_actual.strftime('%d/%m/%Y %H:%M')}"

        headers = ["Apartamento", "Nombre", "Coeficiente", "Fecha Ingreso", "Asistencia", "Poder"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        row_num = 5
        for p in participantes:
            fecha_ingreso = "-"
            if p.get("login_time") and p.get("present"):
                try:
                    login_time_str = str(p["login_time"])
                    if 'T' in login_time_str:
                        dt_utc = datetime.fromisoformat(login_time_str.replace('Z', '+00:00'))
                    else:
                        dt_utc = datetime.strptime(login_time_str, '%Y-%m-%d %H:%M:%S').replace(tzinfo=timezone.utc)
                    fecha_ingreso = dt_utc.astimezone(colombia_tz).strftime('%d/%m/%Y %H:%M')
                except Exception:
                    fecha_ingreso = "Error"

            ws.cell(row=row_num, column=1, value=str(p.get("code", "")))
            ws.cell(row=row_num, column=2, value=str(p.get("name", "")))
            ws.cell(row=row_num, column=3, value=float(p.get("coefficient", 0)))
            ws.cell(row=row_num, column=4, value=fecha_ingreso)
            ws.cell(row=row_num, column=5, value="SI" if p.get("present") else "NO")
            ws.cell(row=row_num, column=6, value=("Si" if p.get("is_power") else "No") if p.get("present") else "-")
            row_num += 1

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=asistencia_{conjunto_name.replace(' ', '_')}.xlsx"}
        )
    except Exception as e:
        logger.error(f"Error creando Excel: {e}")
        raise HTTPException(status_code=500, detail=f"Error creando Excel: {str(e)}")

@router.get("/check/{code}")
def check_participant_exists(code: str):
    conn = get_db()
    try:
        participant = execute_query(conn, "SELECT present FROM participants WHERE code = ?", (code,), fetchone=True)
        return {"exists": participant is not None and participant["present"] == 1}
    finally:
        close_db(conn)

@router.get("/info/{code}")
def get_participant_info(code: str):
    conn = get_db()
    try:
        participant = execute_query(
            conn,
            "SELECT code, name, coefficient FROM participants WHERE code = ?",
            (code,),
            fetchone=True
        )
        if not participant:
            raise HTTPException(status_code=404, detail="Participante no encontrado")
        return dict(participant)
    finally:
        close_db(conn)